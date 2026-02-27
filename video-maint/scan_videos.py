#!/usr/bin/env python3
"""Scan video files for resolution and output to Excel spreadsheet."""
# /// script
# requires-python = ">=3.12"
# dependencies = ["openpyxl"]
# ///

import json
import os
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

VIDEO_EXTENSIONS = {
    ".mp4", ".mkv", ".avi", ".mov", ".wmv", ".flv",
    ".webm", ".m4v", ".ts", ".mpg", ".mpeg",
}
SCAN_DIR = Path("/Volumes/RAID Store/Downloads")
WIDTH_THRESHOLD = 1920  # Pre-check "delete" for width < this
OUTPUT_FILE = Path("/Users/oshyan/Projects/Coding/small/video-maint/video_library.xlsx")


def find_video_files(root: Path) -> list[Path]:
    videos = []
    for dirpath, _, filenames in os.walk(root):
        for f in filenames:
            if Path(f).suffix.lower() in VIDEO_EXTENSIONS:
                videos.append(Path(dirpath) / f)
    videos.sort(key=lambda p: p.name.lower())
    return videos


def get_resolution(filepath: Path) -> tuple[int | None, int | None]:
    """Use ffprobe to get width and height of the first video stream."""
    try:
        result = subprocess.run(
            [
                "ffprobe", "-v", "error",
                "-select_streams", "v:0",
                "-show_entries", "stream=width,height",
                "-of", "json",
                str(filepath),
            ],
            capture_output=True,
            text=True,
            timeout=30,
        )
        data = json.loads(result.stdout)
        streams = data.get("streams", [])
        if streams:
            w = streams[0].get("width")
            h = streams[0].get("height")
            return (w, h)
    except (subprocess.TimeoutExpired, json.JSONDecodeError, Exception) as e:
        print(f"  Error probing {filepath.name}: {e}", file=sys.stderr)
    return (None, None)


def write_spreadsheet(records: list[dict], output: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Video Library"

    headers = ["Filename", "Width", "Height", "Resolution", "Delete", "Full Path"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Pre-check fill for rows marked delete
    delete_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")

    for row_idx, rec in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=rec["filename"])
        ws.cell(row=row_idx, column=2, value=rec["width"])
        ws.cell(row=row_idx, column=3, value=rec["height"])
        ws.cell(row=row_idx, column=4, value=rec["resolution"])
        delete_val = "YES" if rec["delete"] else ""
        ws.cell(row=row_idx, column=5, value=delete_val)
        ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=6, value=rec["path"])

        if rec["delete"]:
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).fill = delete_fill

    # Column widths
    ws.column_dimensions["A"].width = 60  # Filename
    ws.column_dimensions["B"].width = 8   # Width
    ws.column_dimensions["C"].width = 8   # Height
    ws.column_dimensions["D"].width = 14  # Resolution
    ws.column_dimensions["E"].width = 8   # Delete
    ws.column_dimensions["F"].width = 100 # Full Path

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:F{len(records) + 1}"

    wb.save(output)


def main():
    print(f"Scanning: {SCAN_DIR}")
    print(f"Threshold: width < {WIDTH_THRESHOLD}px -> pre-checked for deletion")
    print()

    videos = find_video_files(SCAN_DIR)
    total = len(videos)
    print(f"Found {total} video files. Starting resolution scan...")
    print()

    records = []
    errors = 0
    start = time.time()

    for i, vpath in enumerate(videos, 1):
        w, h = get_resolution(vpath)

        if w is not None and h is not None:
            resolution = f"{w}x{h}"
            should_delete = w < WIDTH_THRESHOLD
        else:
            resolution = "ERROR"
            should_delete = False
            errors += 1

        records.append({
            "filename": vpath.name,
            "width": w,
            "height": h,
            "resolution": resolution,
            "delete": should_delete,
            "path": str(vpath),
        })

        # Progress every 50 files
        if i % 50 == 0 or i == total:
            elapsed = time.time() - start
            rate = i / elapsed if elapsed > 0 else 0
            eta = (total - i) / rate if rate > 0 else 0
            mins, secs = divmod(int(eta), 60)
            print(f"  [{i}/{total}] {rate:.1f} files/sec, ETA: {mins}m{secs:02d}s")

    elapsed = time.time() - start
    mins, secs = divmod(int(elapsed), 60)
    print()
    print(f"Scan complete in {mins}m{secs:02d}s")

    delete_count = sum(1 for r in records if r["delete"])
    keep_count = total - delete_count - errors
    print(f"  Total files:     {total}")
    print(f"  Pre-checked DEL: {delete_count} (width < {WIDTH_THRESHOLD})")
    print(f"  Keep:            {keep_count}")
    print(f"  Errors:          {errors}")

    write_spreadsheet(records, OUTPUT_FILE)
    print(f"\nSpreadsheet written to: {OUTPUT_FILE}")
    print("Open it, review the 'Delete' column (YES = pre-checked), then save when done.")


if __name__ == "__main__":
    main()
