#!/usr/bin/env python3
"""Convert the library CSV to a formatted Excel workbook."""
# /// script
# requires-python = ">=3.12"
# dependencies = ["openpyxl"]
# ///

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

BASE_DIR = Path(__file__).resolve().parent
INPUT_CSV = BASE_DIR / "video_library.csv"
OUTPUT_XLSX = BASE_DIR / "video_library.xlsx"

VIDEO_EXTENSIONS = {
    ".mp4", ".mkv", ".avi", ".mov", ".wmv", ".flv",
    ".webm", ".m4v", ".ts", ".mpg", ".mpeg",
}
HEADERS = [
    "Filename",
    "Category",
    "Extension",
    "Size Bytes",
    "Size",
    "Width",
    "Height",
    "Resolution",
    "Delete",
    "Full Path",
]


def extension_for(filename, path):
    source = filename or path or ""
    return Path(source).suffix.lower()


def infer_category(row):
    category = row.get("Category", "").strip()
    if category:
        return category
    if extension_for(row.get("Filename", ""), row.get("Full Path", "")) in VIDEO_EXTENSIONS:
        return "Video"
    if row.get("Width") or row.get("Height") or row.get("Resolution"):
        return "Video"
    return "Other"


def canonical_row(row):
    filename = row.get("Filename", "")
    path = row.get("Full Path", row.get("Path", ""))
    extension = row.get("Extension", "") or extension_for(filename, path)
    category = infer_category(row)
    size_bytes = row.get("Size Bytes", row.get("SizeBytes", ""))
    return {
        "Filename": filename,
        "Category": category,
        "Extension": extension,
        "Size Bytes": int(size_bytes) if str(size_bytes).isdigit() else None,
        "Size": row.get("Size", ""),
        "Width": int(row["Width"]) if row.get("Width") else None,
        "Height": int(row["Height"]) if row.get("Height") else None,
        "Resolution": row.get("Resolution", ""),
        "Delete": row.get("Delete", ""),
        "Full Path": path,
    }


def add_sheet(wb, title, rows):
    ws = wb.create_sheet(title)
    header_fill = PatternFill(start_color="26352D", end_color="26352D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    delete_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    error_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, header in enumerate(HEADERS, 1):
            value = row[header]
            ws.cell(row=row_idx, column=col_idx, value=value)

        ws.cell(row=row_idx, column=9).alignment = Alignment(horizontal="center")

        if row["Delete"] == "YES":
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=col).fill = delete_fill
        elif row["Resolution"] == "ERROR":
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=col).fill = error_fill

    widths = {
        "A": 65,
        "B": 12,
        "C": 12,
        "D": 13,
        "E": 12,
        "F": 8,
        "G": 8,
        "H": 14,
        "I": 8,
        "J": 110,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{max(ws.max_row, 1)}"
    return ws


def main():
    with open(INPUT_CSV, newline="", encoding="utf-8-sig") as f:
        rows = [canonical_row(row) for row in csv.DictReader(f)]

    wb = Workbook()
    del wb[wb.active.title]

    videos = [row for row in rows if row["Category"] == "Video"]
    others = [row for row in rows if row["Category"] != "Video"]
    add_sheet(wb, "Videos", videos)
    add_sheet(wb, "Other Files", others)

    wb.save(OUTPUT_XLSX)
    print(f"Written: {OUTPUT_XLSX}")
    print(f"Videos: {len(videos)}")
    print(f"Other files: {len(others)}")


if __name__ == "__main__":
    main()
